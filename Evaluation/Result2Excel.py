import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def export_scores_to_excel(scores, etype, include_turn_acc=True, file_name="evaluation_results.xlsx"):
    """
    将评分数据导出到Excel文件。

    Args:
        scores (dict): 包含各难度级别和类型的评分数据。
        etype (str): 要导出的评分类型，可选值为 "exec"（执行准确率）、"match"（完全匹配准确率）或 "all"（两者都包含）。
        include_turn_acc (bool, optional): 是否包含回合准确率，默认为True。
        file_name (str, optional): 导出Excel文件的名称，默认为 "evaluation_results.xlsx"。

    Returns:
        None

    Raises:
        None

    """
    turns = ['turn 1', 'turn 2', 'turn 3', 'turn 4', 'turn > 4']
    levels = ['easy', 'medium', 'hard', 'extra', 'all']
    if include_turn_acc:
        levels.append('joint_all')
    partial_types = ['select', 'select(no AGG)', 'where', 'where(no OP)', 'group(no Having)',
                     'group', 'order', 'and/or', 'IUEN', 'keywords']

    rows = []

    # 添加 count 行
    rows.append(['count'] + [scores[level]['count'] for level in levels])

    # EXECUTION ACCURACY
    if etype in ["all", "exec"]:
        rows.append(['------------------------------   EXECUTION ACCURACY     ------------------------------'])
        rows.append(['execution'] + [round(float(scores[level]['exec']), 3) for level in levels])

    # EXACT MATCHING ACCURACY
    if etype in ["all", "match"]:
        rows.append(['------------------------------ EXACT MATCHING ACCURACY  ------------------------------'])
        rows.append(['exact match'] + [round(float(scores[level]['exact']), 3) for level in levels])

        # PARTIAL MATCHING ACCURACY
        rows.append(['------------------------------PARTIAL MATCHING ACCURACY-------------------------------'])
        for type_ in partial_types:
            rows.append([type_] + [round(float(scores[level]['partial'][type_]['acc']), 3) for level in levels])

        # PARTIAL MATCHING RECALL
        rows.append(['------------------------------- PARTIAL MATCHING RECALL -------------------------------'])
        for type_ in partial_types:
            rows.append([type_] + [round(float(scores[level]['partial'][type_]['rec']), 3) for level in levels])

        # PARTIAL MATCHING F1
        rows.append(['------------------------------- PARTIAL MATCHING F1 -----------------------------------'])
        for type_ in partial_types:
            rows.append([type_] + [round(float(scores[level]['partial'][type_]['f1']), 3) for level in levels])

    # 将数据转换为 DataFrame 并导出到 Excel
    df = pd.DataFrame(rows)
    df.to_excel(file_name, index=False, header=[''] + levels)

    # 打开生成的Excel文件并设置格式
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    df.to_excel(writer, index=False, header=[''] + levels)
    writer.close()

    # 打开生成的Excel文件进行格式化
    workbook = load_workbook(file_name)
    worksheet = workbook.active

    # 合并指定的标题行并居中
    for row_idx in range(len(rows)):
        if "---------------------" in rows[row_idx][0]:
            worksheet.merge_cells(start_row=row_idx + 2, start_column=1, end_row=row_idx + 2,
                                  end_column=len(levels) + 1)
            worksheet.cell(row=row_idx + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 设置所有单元格居中
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置每一列的宽度为20
    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = 20  # 大约对应250像素

    # 设置行高为17
    for row in worksheet.iter_rows():
        worksheet.row_dimensions[row[0].row].height = 17

    # 保存格式化的文件
    workbook.save(file_name)
    print(f"评估结果已成功保存到 {file_name}")
